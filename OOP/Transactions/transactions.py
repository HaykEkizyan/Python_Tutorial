import openpyxl as xl
from openpyxl.chart import BarChart, Reference

def process_workbook(filename):             # for useful we put our all code into function
    wb = xl.load_workbook(filename)         # load workbook
    sheet = wb['Sheet1']                    # reference to the worksheet

    for row in range(2, sheet.max_row + 1):
        cell = sheet.cell(row, 3)
        corrected_price = cell.value * 0.9          # we need to get 90% of our current prices to the D column
        corrected_price_cell = sheet.cell(row, 4)   # D column's number is 4
        corrected_price_cell.value = corrected_price

                                               # use the Reference class to select a range for loop
    values = Reference(sheet,                  # adding sheet argument
                       min_row=2,
                       max_row=sheet.max_row,  # maximum row in this sheet
                       min_col=4,
                       max_col=8)              # limiting the range of cells we're selecting to the fourth column

    chart = BarChart()                         # creating barchart class's instance
    chart.add_data(values)                     # pass our valies to chart.add_data
    sheet.add_chart(chart, 'e2')               # call sheet, pass this chart object, and specify where we want to add this chart

    wb.save(filename)                          # we created a new file in our directory


