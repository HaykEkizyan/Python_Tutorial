import openpyxl as xl                                 # import openpyxl lets us work with Pypi libraries, which include also spreadsheet tools
wb = xl.load_workbook('transactions.xlsx')            # transactions.xlsx file is with app.py file in the current project
sheet = wb['Sheet1']
cell = sheet['a1']
cell = sheet.cell(2, 1)

# print(cell.value) -> returns 1001, which is in A2 line at our file

# print(sheet.max_row) -> returns 4, which shows how many rows we have


